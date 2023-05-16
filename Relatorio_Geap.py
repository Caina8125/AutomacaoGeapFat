import pandas as pd
import pyautogui
import time
from abc import ABC
from tkinter import filedialog
from selenium import webdriver
from openpyxl import Workbook, load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


class PageElement(ABC):
    def __init__(self, webdriver, url=''):
        self.webdriver = webdriver
        self.url = url
    def open(self):
        self.webdriver.get(self.url)

class Login(PageElement):
    multiusuario = (By.XPATH, '/html/body/div[3]/div[3]/div/form/div[1]/label')
    prestador = (By.XPATH, '//*[@id="login_code"]')
    cpf = (By.XPATH, '//*[@id="login_cpf"]')
    senha = (By.XPATH, '//*[@id="login_password"]')
    logar = (By.XPATH, '//*[@id="btnLogin"]')

    def exe_login(self, prestador, cpf, senha):
        self.webdriver.find_element(*self.multiusuario).click()
        self.webdriver.find_element(*self.prestador).send_keys(prestador)
        self.webdriver.find_element(*self.cpf).send_keys(cpf)
        self.webdriver.find_element(*self.senha).send_keys(senha)
        self.webdriver.find_element(*self.logar).click()
        time.sleep(4)

class caminho(PageElement):
    Alerta = (By.XPATH, '/html/body/div[2]/div/center/a')

    def exe_caminho(self):
        try:
            self.webdriver.find_element(*self.Alerta).click()
        except:
            print('Não tem alerta')

        webdriver.get("https://www2.geap.com.br/PRESTADOR/tiss-baixa.asp")
        time.sleep(3)

class capturar_protocolo(PageElement):
    inserir_lote = (By.XPATH, '//*[@id="NroLotePrestador"]')
    baixar = (By.XPATH, '//*[@id="main"]/div/div/div[2]/div[2]/article/form/div/a')
    elemento2 = (By.XPATH, '//*[@id="main"]/div/div/div/table/tbody/tr[2]/td[5]')
    elemento3 = (By.XPATH, '//*[@id="main"]/div/div/div/table/tbody/tr[3]/td[5]')
    

    def exe_capturar(self):

        count = 0
        faturas_df = pd.read_excel(planilha)
        for index, linha in faturas_df.iterrows():

            count = count + 1
            
            fatura =  f"{linha['Fatura']}".replace(".0","")
            self.webdriver.find_element(*self.inserir_lote).send_keys(fatura)
            time.sleep(1)
            self.webdriver.find_element(*self.baixar).click()
            time.sleep(1)
            try:
                erro_fatura = WebDriverWait(webdriver, 10).until(EC.presence_of_element_located((By.XPATH,'//*[@id="main"]/div/div/div/div')))
                fatura_erro = ["Fatura inexistente"]
                df = pd.DataFrame(fatura_erro)
                book = load_workbook(planilha)
                writer = pd.ExcelWriter(planilha, engine='openpyxl')
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                df.to_excel(writer, 'Fatura', startrow= count, startcol=1, header=False, index=False)
                writer.save()
                time.sleep(2)
                webdriver.get("https://www2.geap.com.br/PRESTADOR/tiss-baixa.asp")
                continue                   
            except:
                print("Fatura correta")

            try:
                verifica_arquivo1 = webdriver.find_element(By.XPATH, '//*[@id="main"]/div/div/div/table/tbody/tr[2]/td[9]').text
            except:
                None

            #try:
                #verifica_arquivo2 = webdriver.find_element(By.XPATH, '//*[@id="main"]/div/div/div/table/tbody/tr[3]/td[9]').text
            #except:
                #None

            if verifica_arquivo1 == "---":
                print("Primeiro arquivo não passou! Buscando o segundo...")
                protocolo = webdriver.find_element(By.XPATH, '//*[@id="main"]/div/div/div/table/tbody/tr[3]/td[1]').text
                n_protocolo = [protocolo]
                df = pd.DataFrame(n_protocolo)
                book = load_workbook(planilha)
                writer = pd.ExcelWriter(planilha, engine='openpyxl')
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                df.to_excel(writer, 'Fatura', startrow= count, startcol=1, header=False, index=False)
                writer.save()
                webdriver.get("https://www2.geap.com.br/PRESTADOR/tiss-baixa.asp")

            else:
                protocolo = webdriver.find_element(By.XPATH, '//*[@id="main"]/div/div/div/table/tbody/tr[2]/td[1]').text
                n_protocolo = [protocolo]
                df = pd.DataFrame(n_protocolo)
                book = load_workbook(planilha)
                writer = pd.ExcelWriter(planilha, engine='openpyxl')
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                df.to_excel(writer, 'Fatura', startrow= count, startcol=1, header=False, index=False)
                writer.save()
                print(f"{linha['Fatura']}", "-", f"{linha['Protocolo']}")
                webdriver.get("https://www2.geap.com.br/PRESTADOR/tiss-baixa.asp")

    def auditoria(self):
        count = 0
        faturas_df = pd.read_excel(planilha)
        for index, linha in faturas_df.iterrows():
            count = count + 1 

            if (f"{linha['Protocolo']}") == "Fatura inexistente":
                continue
            else:
                None

            protocolo = f"{linha['Protocolo']}".replace(".0","")
            webdriver.get("https://www2.geap.com.br/PRESTADOR/tiss-capa-de-lote.asp?NroProtocolo=" + protocolo)
            Dados = webdriver.find_element(By.XPATH, '//*[@id="main"]/div/div/div/table/tbody').text
            print(Dados)

            if Dados.count("*") == 0:
                n_contem = ["Não Contém devolução"]
                df = pd.DataFrame(n_contem)
                book = load_workbook(planilha)
                writer = pd.ExcelWriter(planilha, engine='openpyxl')
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                df.to_excel(writer, 'Fatura', startrow= count, startcol=2, header=False, index=False)
                writer.save()
            else:
                contem = ["Contém devolução"]
                df = pd.DataFrame(contem)
                book = load_workbook(planilha)
                writer = pd.ExcelWriter(planilha, engine='openpyxl')
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                df.to_excel(writer, 'Fatura', startrow= count, startcol=2, header=False, index=False)
                writer.save()


#-------------------------------------------------------------------------------------------------------------------------

login_usuario = "lucas.timoteo"
senha_usuario = "Caina9018"

planilha = filedialog.askopenfilename()

url = 'https://www2.geap.com.br/auth/prestador.asp'

webdriver = webdriver.Chrome(r"\\10.0.0.71\atualiza\teste-python\chromedriver.exe")

login_page = Login(webdriver, url)
login_page.open()
webdriver.maximize_window()
time.sleep(4)
pyautogui.write(login_usuario)
pyautogui.press("TAB")
time.sleep(1)
pyautogui.write(senha_usuario)
pyautogui.press("enter")
time.sleep(4)

login_page.exe_login(
    prestador = "23003723",
    cpf = '66661692120',
    senha = "amhpdf0073"
)

time.sleep(4)

caminho(webdriver,url).exe_caminho()

capturar_protocolo(webdriver, url).exe_capturar()

capturar_protocolo(webdriver, url).auditoria()